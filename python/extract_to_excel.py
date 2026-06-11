"""
extract_to_excel.py  —  ZONA | AdventureWorks Data Lakehouse
Step 0 : Source Extraction

Ambil data penjualan TOKO dari database AdventureWorks:
  - Filter : OnlineOrderFlag = false/0  (penjualan toko via salesperson)
  - Join   : SalesOrderHeader + SalesOrderDetail + Product +
             ProductSubcategory + ProductCategory + SalesPerson + SalesTerritory
  - Output : datalake/bronze/store_sales/store_sales_raw.xlsx

Mendukung dua mode:
  - Mode DUMP  : baca dari file .tar PostgreSQL dump (offline)
  - Mode LIVE  : connect langsung ke PostgreSQL
"""

import sys
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, str(Path(__file__).parent))
from config import DUMP_PATH, POSTGRESQL, EXCEL_FILE, BRONZE_DIR


# ────────────────────────────────────────────────────────────────
#  Mode 1: Baca dari PostgreSQL dump file (.tar)
# ────────────────────────────────────────────────────────────────
def extract_from_dump() -> pd.DataFrame:
    import pgdumplib
    print(f"[DUMP]  Membaca {DUMP_PATH.name}...")
    dump = pgdumplib.load(str(DUMP_PATH))

    # ── salesorderheader ──────────────────────────────────────
    hdr_cols = [
        'salesorderid','revisionnumber','orderdate','duedate','shipdate',
        'status','onlineorderflag','purchaseordernumber','accountnumber',
        'customerid','salespersonid','territoryid','billtoaddressid',
        'shiptoaddressid','shipmethodid','creditcardid','creditcardapprovalcode',
        'currencyrateid','subtotal','taxamt','freight','totaldue',
        'comment','rowguid_hdr','modifieddate_hdr'
    ]
    df_hdr = pd.DataFrame(
        list(dump.table_data('sales', 'salesorderheader')),
        columns=hdr_cols
    )
    # Filter store sales (OnlineOrderFlag = false)
    df_hdr = df_hdr[df_hdr['onlineorderflag'] == 'f'].copy()
    print(f"[DUMP]  salesorderheader (store only): {len(df_hdr)} orders")

    # ── salesorderdetail ──────────────────────────────────────
    det_cols = [
        'salesorderid','salesorderdetailid','carriertrackingnumber',
        'orderqty','productid','specialofferid','unitprice',
        'unitpricediscount','rowguid_det','modifieddate_det'
    ]
    df_det = pd.DataFrame(
        list(dump.table_data('sales', 'salesorderdetail')),
        columns=det_cols
    )
    print(f"[DUMP]  salesorderdetail: {len(df_det)} baris")

    # ── product ───────────────────────────────────────────────
    prod_cols = [
        'productid','name','productnumber','makeflag','finishedgoodsflag',
        'color','safetystocklevel','reorderpoint','standardcost','listprice',
        'size','sizeunitmeasurecode','weightunitmeasurecode','weight',
        'daystomanufacture','productline','class','style',
        'productsubcategoryid','productmodelid','sellstartdate','sellenddate',
        'discontinueddate','rowguid_prod','modifieddate_prod'
    ]
    df_prod = pd.DataFrame(
        list(dump.table_data('production', 'product')),
        columns=prod_cols
    )[['productid','name','productsubcategoryid','listprice','productline','color']]
    df_prod.rename(columns={'name': 'productname'}, inplace=True)

    # ── productsubcategory ────────────────────────────────────
    psub_cols = ['productsubcategoryid','productcategoryid','name','rowguid_ps','modifieddate_ps']
    df_psub = pd.DataFrame(
        list(dump.table_data('production', 'productsubcategory')),
        columns=psub_cols
    )[['productsubcategoryid','productcategoryid','name']]
    df_psub.rename(columns={'name': 'subcategoryname'}, inplace=True)

    # ── productcategory ───────────────────────────────────────
    pcat_cols = ['productcategoryid','name','rowguid_pc','modifieddate_pc']
    df_pcat = pd.DataFrame(
        list(dump.table_data('production', 'productcategory')),
        columns=pcat_cols
    )[['productcategoryid','name']]
    df_pcat.rename(columns={'name': 'categoryname'}, inplace=True)

    # ── salesperson ───────────────────────────────────────────
    sp_cols = [
        'businessentityid','territoryid_sp','salesquota','bonus',
        'commissionpct','salesytd','saleslastyear','rowguid_sp','modifieddate_sp'
    ]
    df_sp = pd.DataFrame(
        list(dump.table_data('sales', 'salesperson')),
        columns=sp_cols
    )[['businessentityid','territoryid_sp','bonus','commissionpct','salesytd']]
    df_sp.rename(columns={'businessentityid': 'salespersonid'}, inplace=True)

    # ── salesterritory ────────────────────────────────────────
    ter_cols = [
        'territoryid','name','countryregioncode','group',
        'salesytd_ter','saleslastyear_ter','costytd','costlastyear',
        'rowguid_ter','modifieddate_ter'
    ]
    df_ter = pd.DataFrame(
        list(dump.table_data('sales', 'salesterritory')),
        columns=ter_cols
    )[['territoryid','name','countryregioncode','group']]
    df_ter.rename(columns={
        'name': 'territoryname',
        'group': 'territorygroup'
    }, inplace=True)

    # ── JOIN semua ────────────────────────────────────────────
    # 1. Header (store only) + Detail
    df = df_hdr.merge(df_det, on='salesorderid', how='inner')

    # 2. + Product
    df = df.merge(df_prod, on='productid', how='left')

    # 3. + Subcategory & Category
    df = df.merge(df_psub, on='productsubcategoryid', how='left')
    df = df.merge(df_pcat, on='productcategoryid', how='left')

    # 4. + Salesperson
    df['salespersonid'] = df['salespersonid'].astype(str)
    df_sp['salespersonid'] = df_sp['salespersonid'].astype(str)
    df = df.merge(df_sp, on='salespersonid', how='left')

    # 5. + Territory
    df['territoryid'] = df['territoryid'].astype(str)
    df_ter['territoryid'] = df_ter['territoryid'].astype(str)
    df = df.merge(df_ter, on='territoryid', how='left')

    # ── Pilih & rename kolom final ────────────────────────────
    final_cols = {
        'salesorderid'         : 'SalesOrderID',
        'salesorderdetailid'   : 'SalesOrderDetailID',
        'orderdate'            : 'OrderDate',
        'duedate'              : 'DueDate',
        'shipdate'             : 'ShipDate',
        'status'               : 'Status',
        'onlineorderflag'      : 'OnlineOrderFlag',
        'purchaseordernumber'  : 'PurchaseOrderNumber',
        'customerid'           : 'CustomerID',
        'salespersonid'        : 'SalespersonID',
        'territoryid'          : 'TerritoryID',
        'territoryname'        : 'TerritoryName',
        'countryregioncode'    : 'TerritoryCountry',
        'territorygroup'       : 'TerritoryGroup',
        'subtotal'             : 'HeaderSubTotal',
        'taxamt'               : 'TaxAmt',
        'freight'              : 'Freight',
        'totaldue'             : 'TotalDue',
        'productid'            : 'ProductID',
        'productname'          : 'ProductName',
        'subcategoryname'      : 'ProductSubcategory',
        'categoryname'         : 'ProductCategory',
        'listprice'            : 'ListPrice',
        'productline'          : 'ProductLine',
        'color'                : 'Color',
        'orderqty'             : 'OrderQty',
        'unitprice'            : 'UnitPrice',
        'unitpricediscount'    : 'UnitPriceDiscount',
        'bonus'                : 'SalespersonBonus',
        'commissionpct'        : 'CommissionPct',
        'salesytd'             : 'SalesYTD',
    }
    existing = {k: v for k, v in final_cols.items() if k in df.columns}
    df = df[list(existing.keys())].rename(columns=existing)

    print(f"[DUMP]  Final joined rows: {len(df)}")
    return df


# ────────────────────────────────────────────────────────────────
#  Mode 2: Connect langsung ke PostgreSQL
# ────────────────────────────────────────────────────────────────
LIVE_QUERY = """
SELECT
    soh.salesorderid            AS "SalesOrderID",
    sod.salesorderdetailid      AS "SalesOrderDetailID",
    soh.orderdate               AS "OrderDate",
    soh.duedate                 AS "DueDate",
    soh.shipdate                AS "ShipDate",
    soh.status                  AS "Status",
    soh.onlineorderflag         AS "OnlineOrderFlag",
    soh.purchaseordernumber     AS "PurchaseOrderNumber",
    soh.customerid              AS "CustomerID",
    soh.salespersonid           AS "SalespersonID",
    soh.territoryid             AS "TerritoryID",
    ter.name                    AS "TerritoryName",
    ter.countryregioncode       AS "TerritoryCountry",
    ter.\"group\"               AS "TerritoryGroup",
    soh.subtotal                AS "HeaderSubTotal",
    soh.taxamt                  AS "TaxAmt",
    soh.freight                 AS "Freight",
    soh.totaldue                AS "TotalDue",
    sod.productid               AS "ProductID",
    p.name                      AS "ProductName",
    ps.name                     AS "ProductSubcategory",
    pc.name                     AS "ProductCategory",
    p.listprice                 AS "ListPrice",
    p.productline               AS "ProductLine",
    p.color                     AS "Color",
    sod.orderqty                AS "OrderQty",
    sod.unitprice               AS "UnitPrice",
    sod.unitpricediscount       AS "UnitPriceDiscount",
    sp.bonus                    AS "SalespersonBonus",
    sp.commissionpct            AS "CommissionPct",
    sp.salesytd                 AS "SalesYTD"
FROM   sales.salesorderheader  soh
JOIN   sales.salesorderdetail  sod ON sod.salesorderid = soh.salesorderid
JOIN   production.product      p   ON p.productid = sod.productid
LEFT JOIN production.productsubcategory ps  ON ps.productsubcategoryid = p.productsubcategoryid
LEFT JOIN production.productcategory    pc  ON pc.productcategoryid = ps.productcategoryid
LEFT JOIN sales.salesperson             sp  ON sp.businessentityid = soh.salespersonid
LEFT JOIN sales.salesterritory          ter ON ter.territoryid = soh.territoryid
WHERE  soh.onlineorderflag = false   -- *** FILTER: penjualan TOKO saja ***
ORDER  BY soh.salesorderid, sod.salesorderdetailid
"""


def extract_from_live() -> pd.DataFrame:
    from sqlalchemy import create_engine, text
    cfg = POSTGRESQL
    url = (f"postgresql+psycopg2://{cfg['user']}:{cfg['password']}"
           f"@{cfg['host']}:{cfg['port']}/{cfg['database']}")
    engine = create_engine(url)
    print(f"[LIVE]  Connecting to PostgreSQL {cfg['database']}...")
    with engine.connect() as conn:
        df = pd.read_sql_query(text(LIVE_QUERY), conn)
    print(f"[LIVE]  Query selesai: {len(df)} baris")
    return df


# ────────────────────────────────────────────────────────────────
#  Style Excel
# ────────────────────────────────────────────────────────────────
def style_excel(path: Path):
    wb = load_workbook(path)
    ws = wb.active
    ws.title = "StoreSales"

    fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = align

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


# ────────────────────────────────────────────────────────────────
#  Main
# ────────────────────────────────────────────────────────────────
def extract():
    print("=" * 60)
    print("  STEP 0 — Extract AdventureWorks → Excel")
    print("  Filter: OnlineOrderFlag = 0  (Store Sales)")
    print("=" * 60)

    # Pilih mode: coba live dulu, fallback ke dump
    df = None
    if POSTGRESQL.get("password") != "" or POSTGRESQL.get("host") != "localhost":
        try:
            df = extract_from_live()
        except Exception as e:
            print(f"[WARN]  Live DB gagal ({e}), fallback ke dump file.")

    if df is None:
        if not DUMP_PATH.exists():
            raise FileNotFoundError(
                f"Dump file tidak ada: {DUMP_PATH}\n"
                f"Letakkan database.tar di folder root project."
            )
        df = extract_from_dump()

    if len(df) == 0:
        print("⚠️  Tidak ada data store sales (OnlineOrderFlag=0)")
        return df

    # Tambah metadata ekstraksi
    df["_extracted_at"] = datetime.now().isoformat()
    df["_source"]       = "AdventureWorks (OnlineOrderFlag=0)"

    # Simpan ke Excel
    BRONZE_DIR.mkdir(parents=True, exist_ok=True)
    print(f"\n[EXCEL] Menyimpan {len(df):,} baris ke {EXCEL_FILE.name}...")
    df.to_excel(EXCEL_FILE, index=False, sheet_name="StoreSales", engine="openpyxl")
    style_excel(EXCEL_FILE)

    size_mb = EXCEL_FILE.stat().st_size / (1024**2)
    print(f"[EXCEL] ✅ Tersimpan — {size_mb:.1f} MB")
    print(f"\n  Ringkasan data:")
    print(f"  • Total rows    : {len(df):,}")
    print(f"  • Unique orders : {df['SalesOrderID'].nunique():,}")
    print(f"  • Salesperson   : {df['SalespersonID'].nunique():,}")
    print(f"  • Territory     : {df['TerritoryName'].nunique() if 'TerritoryName' in df else 'N/A'}")
    if 'OrderDate' in df.columns:
        print(f"  • Periode       : {df['OrderDate'].min()} s/d {df['OrderDate'].max()}")
    print(f"\n✅ Extract selesai! Lanjut: python bronze_excel.py")
    return df


if __name__ == "__main__":
    extract()
